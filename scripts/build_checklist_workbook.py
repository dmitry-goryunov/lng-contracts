"""Rebuild LNG_SPA_value_item_checklist.xlsx from checklist_data.py (single source of truth).

Run from the repo root:  python scripts/build_checklist_workbook.py
Requires: openpyxl. Optional: recalculate afterwards with LibreOffice if formulas matter.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from checklist_data import R, ROWS, tt, hg, HOLDER

OUT = Path(__file__).resolve().parents[1] / "LNG_SPA_value_item_checklist.xlsx"

wb = Workbook()
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F4E79")
CAT_FILL = PatternFill("solid", start_color="DEEAF6")
RAG_FILL = {"RED": PatternFill("solid", start_color="F4CCCC"),
            "AMBER": PatternFill("solid", start_color="FFF2CC"),
            "GREEN": PatternFill("solid", start_color="D9EAD3")}
RAG_FONT = {"RED": Font(name=ARIAL, size=9, color="990000", bold=True),
            "AMBER": Font(name=ARIAL, size=9, color="7F6000", bold=True),
            "GREEN": Font(name=ARIAL, size=9, color="274E13", bold=True)}
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN

# ReadMe
ws = wb.active
ws.title = "ReadMe"
readme = [
    ("LNG SPA / GSA value-item checklist", ""),
    ("", ""),
    ("Purpose", "Decompose one LNG SPA into checkable value items: price, start, end, yearly flexibilities, optionalities, plus supporting terms that change their value."),
    ("How to use", "1) Copy the Template sheet, one per contract. 2) Work top to bottom; record values verbatim with clause references. 3) Set Status per row. 4) Compare contracts on the Matrix sheet and RAG-score."),
    ("Status codes", "Extracted = value recorded from the text. Redacted = clause found, number hidden. Not found = searched, absent. Not reviewed = not yet checked. N/A = structurally inapplicable."),
    ("RAG codes", "RED = visible economic difference. AMBER = similar framework, unresolved caveats. GREEN = no material difference detected. ACQ is a scale input, never RAG-scored."),
    ("Epistemic rules", "Never guess a redacted number; treat redacted constants as scenario inputs. Distinguish fact / inference / unknown. 'Not reviewed in deck' means the source decks did not cover the item, not that the contract lacks it."),
    ("Sources", "Pre-filled Matrix entries come only from: cheniere_sabine_pass_spa_standalone_economic_terms.pptx, driftwood_lng_spa_economic_terms_comparison_formula_signature_updated.pptx and cheniere_followon_gdf_spa_economic_terms.pptx (this folder). Verify against the filed agreements before relying on any figure."),
    ("Companion file", "LNG_SPA_value_item_checklist.md holds the methodology and the same checklist in portable form."),
    ("Term type", "Cash-flow = feeds the base DCF strip. Option = embedded optionality; requires option-adjusted valuation (Monte Carlo / LSMC / spread-option methods), not static DCF. Risk-structural = changes risk, credit or enforceability rather than the base strip. Context = identity and reference data."),
    ("Hedgeability", "Curve = hedgeable against liquid benchmarks (Brent, HH, TTF, JKM, FX, SOFR). Spread = hedgeable as a location, time or freight spread. Residual = imperfectly hedgeable; goes on the basis-risk register. Tags are indicative defaults; review per contract."),
    ("Enforceability rule", "Before assigning value to any Option row (destination, rejection, deferral, price review), check legal and competition-law enforceability (F13). If exercise is uncertain, haircut or exclude the option value."),
    ("Holder inputs", "The Holder_Inputs sheet lists valuation inputs that are NOT contract terms (freight, regas tariffs, FX, discounting, downstream use value). Keep them separate from extracted contract facts."),
    ("v2 update", "5 Jul 2026: folded in improvements from 'LNG Contract Valuation and Hedging.docx': rows D13, E13, E14, F10-F13, I8 added; G1 enriched with period basis; Term type and Hedgeability columns; Holder_Inputs sheet."),
]
for i, (k, v) in enumerate(readme, 1):
    ws.cell(row=i, column=1, value=k).font = Font(name=ARIAL, size=10, bold=True)
    c = ws.cell(row=i, column=2, value=v)
    c.font = Font(name=ARIAL, size=10)
    c.alignment = WRAP
ws.cell(row=1, column=1).font = Font(name=ARIAL, size=14, bold=True)
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 120

# Template
ws = wb.create_sheet("Template")
tcols = ["ID", "Category", "Value item", "What to record", "Typical clause location", "Why it matters for value", "Term type", "Hedgeability", "Status", "Value found", "Clause ref", "RAG", "Notes"]
ws.cell(row=1, column=1, value="Contract:").font = Font(name=ARIAL, size=10, bold=True)
ws.cell(row=1, column=2, value="<name>").font = Font(name=ARIAL, size=10, color="0000FF")
ws.cell(row=1, column=4, value="Items:").font = Font(name=ARIAL, size=10, bold=True)
ws.cell(row=1, column=5, value=f"=COUNTA(A4:A{3+len(R)})").font = Font(name=ARIAL, size=10)
ws.cell(row=1, column=6, value="Extracted:").font = Font(name=ARIAL, size=10, bold=True)
ws.cell(row=1, column=7, value=f"=COUNTIF(I4:I{3+len(R)},\"Extracted\")").font = Font(name=ARIAL, size=10)
ws.cell(row=1, column=8, value="Open (blank/Not reviewed):").font = Font(name=ARIAL, size=10, bold=True)
ws.cell(row=1, column=9, value=f"=COUNTIF(I4:I{3+len(R)},\"Not reviewed\")+COUNTBLANK(I4:I{3+len(R)})").font = Font(name=ARIAL, size=10)
for j, h in enumerate(tcols, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(tcols))
row = 4
for t in R:
    for j, v in enumerate([t[0], t[1], t[2], t[3], t[4], t[5], tt(t[0]), hg(t[0])], 1):
        ws.cell(row=row, column=j, value=v)
    for c in range(1, 14):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=ARIAL, size=9)
        cell.alignment = WRAP
        cell.border = THIN
    if t[0].endswith("1") and len(t[0]) == 2:
        for c in range(1, 14):
            ws.cell(row=row, column=c).fill = CAT_FILL
    row += 1
dv_status = DataValidation(type="list", formula1='"Extracted,Redacted,Not found,Not reviewed,N/A"', allow_blank=True)
dv_rag = DataValidation(type="list", formula1='"RED,AMBER,GREEN,N/A"', allow_blank=True)
ws.add_data_validation(dv_status); ws.add_data_validation(dv_rag)
dv_status.add(f"I4:I{row-1}"); dv_rag.add(f"L4:L{row-1}")
widths = [6, 22, 26, 28, 18, 32, 12, 11, 12, 30, 11, 9, 22]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "D4"

# Matrix
ws = wb.create_sheet("Matrix_13_contracts")
mcols = ["ID", "Category", "Value item", "Term type", "Hedgeability", "BG (SPL)", "Gas Natural Fenosa (SPL)", "GAIL (SPL)", "Total (SPL)",
         "Vitol (DWL)", "Gunvor (DWL)", "Shell SPA1 JKM (DWL)", "Shell SPA2 TTF (DWL)",
         "BG-GC A&R (SPL)", "Centrica (SPL)", "GNF (CCL)", "Woodside (CCL)", "GdF (Master, DES)",
         "RAG Cheniere set", "RAG Driftwood set"]
ws.cell(row=1, column=1, value="Pre-filled from the three source decks only. 'Not reviewed in deck' = item not covered there. Redacted values shown as [***]; never guess them. Term-type and hedgeability tags are indicative defaults; review per contract. SPL = Sabine Pass Liquefaction; DWL = Driftwood LNG; CCL = Corpus Christi Liquefaction.").font = Font(name=ARIAL, size=9, italic=True)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
for j, h in enumerate(mcols, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(mcols))
row = 4
for t in ROWS:
    vals = ([t[0], t[1], t[2], tt(t[0]), hg(t[0])] + list(t[8:21]) + [t[6] or "-", t[7] or "-"])
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = Font(name=ARIAL, size=9)
        cell.alignment = WRAP
        cell.border = THIN
    for j in (19, 20):
        v = ws.cell(row=row, column=j).value
        if v in RAG_FILL:
            ws.cell(row=row, column=j).fill = RAG_FILL[v]
            ws.cell(row=row, column=j).font = RAG_FONT[v]
    if t[0].endswith("1") and len(t[0]) == 2:
        for j in range(1, 6):
            ws.cell(row=row, column=j).fill = CAT_FILL
    row += 1
mw = [6, 22, 24, 12, 11, 26, 26, 26, 28, 30, 30, 30, 30, 30, 28, 28, 28, 30, 11, 11]
for j, w in enumerate(mw, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "D4"
ws.auto_filter.ref = f"A3:T{row-1}"

# Holder_Inputs
ws = wb.create_sheet("Holder_Inputs")
hcols = ["Input", "What to record", "Typical source", "Why it matters for value"]
ws.cell(row=1, column=1, value="Holder-side valuation inputs. These are NOT contract terms; they complete the valuation identity V = sum D(0,t) x E[(P_use - P_contract) x Q_lift - costs]. Keep them separate from extracted contract facts.").font = Font(name=ARIAL, size=9, italic=True)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
for j, h in enumerate(hcols, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(hcols))
row = 4
for t in HOLDER:
    for j, v in enumerate(t, 1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = Font(name=ARIAL, size=9)
        cell.alignment = WRAP
        cell.border = THIN
    row += 1
for j, w in enumerate([30, 40, 48, 52], 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A4"

wb.save(OUT)
print(f"Saved {OUT} ({len(R)} rows)")
